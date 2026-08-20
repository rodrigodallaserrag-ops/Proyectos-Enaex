"""
Streamlit - Dx Compradores
Réplica del pbix "Nivel_de_servicio_BI.pbix", página "Dx Compradores" y Trazabilidad.

Correr local: streamlit run app.py
"""
import pandas as pd
import streamlit as st

import config
import loaders
import transform

# Importación del archivo ariba_trazabilidad.py con alias para mantener compatibilidad
try:
    import ariba_trazabilidad as trazabilidad
    HAS_TRAZABILIDAD = True
except ImportError:
    HAS_TRAZABILIDAD = False

st.set_page_config(page_title="Dx Compradores - Nivel de Servicio", layout="wide")

# ==============================================================================
# CONFIGURACIÓN TEMA (CLARO PREDETERMINADO / OSCURO)
# ==============================================================================
if "tema" not in st.session_state:
    st.session_state["tema"] = "claro"

# Botón flotante para cambiar tema
icono_tema = "🌙" if st.session_state["tema"] == "claro" else "☀️"
if st.button(icono_tema, key="theme_toggle", help="Alternar Modo Claro/Oscuro"):
    st.session_state["tema"] = "oscuro" if st.session_state["tema"] == "claro" else "claro"
    st.rerun()

if st.session_state["tema"] == "claro":
    # MODO CLARO: CSS ajustado para centrar el icono correctamente
    st.markdown("""
        <style>
        .st-key-theme_toggle {
            position: fixed !important;
            top: 15px !important;
            right: 60px !important;
            z-index: 999999 !important;
            width: auto !important;
        }
        .st-key-theme_toggle button {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            font-size: 1.6rem !important;
            padding: 0 !important;
            margin: 0 !important;
            color: #111111 !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            line-height: 1 !important;
            height: auto !important;
            min-height: unset !important;
        }
        .st-key-theme_toggle button p {
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1 !important;
            font-size: 1.6rem !important;
        }
        .st-key-theme_toggle button:hover {
            transform: scale(1.1) !important;
            background: transparent !important;
        }
        </style>
    """, unsafe_allow_html=True)
else:
    # MODO OSCURO: CSS ajustado para centrar el icono correctamente
    st.markdown("""
        <style>
        .st-key-theme_toggle {
            position: fixed !important;
            top: 15px !important;
            right: 60px !important;
            z-index: 999999 !important;
            width: auto !important;
        }
        .st-key-theme_toggle button {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            font-size: 1.6rem !important;
            padding: 0 !important;
            margin: 0 !important;
            color: #FF3333 !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            line-height: 1 !important;
            height: auto !important;
            min-height: unset !important;
        }
        .st-key-theme_toggle button p {
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1 !important;
            font-size: 1.6rem !important;
        }
        .st-key-theme_toggle button:hover {
            transform: scale(1.1) !important;
            background: transparent !important;
        }

        .stApp, [data-testid="stAppViewContainer"], [data-testid="stSidebar"], html, body, [data-testid="stHeader"] {
            background-color: #0E1117 !important;
            color: #FF3333 !important;
        }

        p, span, label, h1, h2, h3, h4, h5, h6, div, td, th, caption, .stMarkdown {
            color: #FF3333 !important;
        }

        div[data-testid="stButton"] > button:not(.st-key-theme_toggle button) {
            background-color: #CC0000 !important;
            color: #FFFFFF !important;
            border: 1px solid #FF4D4D !important;
            font-weight: bold !important;
        }
        div[data-testid="stButton"] > button:not(.st-key-theme_toggle button):hover {
            background-color: #FF0000 !important;
            color: #FFFFFF !important;
            border-color: #FF6666 !important;
        }

        input, select, textarea, div[data-baseweb="select"] {
            background-color: #1E2329 !important;
            color: #FF3333 !important;
            border-color: #CC0000 !important;
        }

        [data-testid="stForm"], div[data-testid="stVerticalBlock"] > div:has(input[type="password"]) {
            background-color: #1E2329 !important;
            padding: 2rem !important;
            border-radius: 12px !important;
            border: 1px solid #CC0000 !important;
        }
        </style>
    """, unsafe_allow_html=True)


# ---- 0. Función de Clasificación Corregida ----
def determinar_tipo_ariba(row):
    """
    Clasifica las solicitudes garantizando la detección de Ariba No Catalogada:
    - ⚙️ SAP ERP: Serie 1 (100...), Serie 19, CL...
    - ⚪ SAP MRP: Serie 5 (500...) o marca de Solped MRP.
    - 🟢 ARIBA CATALOGADA / DIRECTA: Flujos directos o catalogados con material.
    - 🔵 ARIBA NO CATALOGADA: Serie 6 sin código de material, en Trazabilidad o sin catálogo.
    """
    sol = str(row.get("Solicitud de pedido", "")).strip()
    material = str(row.get("Material", "")).strip()
    tiene_material = bool(material and material.lower() not in ["nan", "none", "n/a", "-", "0", "null"])
    es_mrp_flag = str(row.get("Solped MRP", "")).strip().lower() in ["sí", "si", "true", "mrp", "1"]
    en_trazabilidad = bool(row.get("En_Trazabilidad", False) or row.get("En Trazabilidad", False))

    texto_origen = ""
    for col in ["Tipo_Ariba", "Origen Ariba", "Origen", "Tipo Flujo", "Tipo Pedido"]:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() != "":
            texto_origen += " " + str(row[col]).upper()

    if sol.startswith("5") or es_mrp_flag or "MRP" in texto_origen:
        return "⚪ SAP MRP"

    if sol.startswith("1") or sol.startswith("19") or sol.upper().startswith("CL") or "ERP" in texto_origen:
        return "⚙️ SAP ERP"

    if sol.startswith("6"):
        if "NO CATALOGAD" in texto_origen or "NOCATALOGAD" in texto_origen or "SIN CODIGO" in texto_origen:
            return "🔵 ARIBA NO CATALOGADA"
        elif en_trazabilidad or not tiene_material:
            return "🔵 ARIBA NO CATALOGADA"
        else:
            return "🟢 ARIBA CATALOGADA / DIRECTA"

    if "NO CATALOGAD" in texto_origen or "NOCATALOGAD" in texto_origen:
        return "🔵 ARIBA NO CATALOGADA"
    elif "DIRECTA" in texto_origen or "CATALOGAD" in texto_origen:
        return "🟢 ARIBA CATALOGADA / DIRECTA"

    return "⚪ OTROS"


# ---- Acceso con contraseña ----
if "app_password" in st.secrets:
    if not st.session_state.get("_autenticado"):
        st.title("Dx Compradores — Nivel de Servicio")
        clave_ingresada = st.text_input("Contraseña de acceso", type="password")
        if st.button("Ingresar"):
            if clave_ingresada == st.secrets["app_password"]:
                st.session_state["_autenticado"] = True
                loaders._descargar_onedrive.clear()
                st.session_state.pop("_clave_pipeline", None)
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
        st.stop()

# ---- Pestañas Principales ----
tab_dx, tab_trazabilidad = st.tabs(["📊 Dx Compradores", "🔗 Trazabilidad No Catalogadas"])

# ==============================================================================
# PESTAÑA 1: DX COMPRADORES
# ==============================================================================
with tab_dx:
    st.title("Dx Compradores — Nivel de Servicio")

    # ---- Fuente de datos ----
    with st.sidebar:
        st.header("Datos de entrada")
        modo = st.radio(
            "Origen de datos",
            ["OneDrive (automático)", "Subir archivos", "Archivos locales (data/)"],
            index=0,
        )

        archivo_data = archivo_resp_grupo = archivo_centro = archivo_mrp = None
        if modo == "OneDrive (automático)":
            archivo_data = "onedrive:me5a_parquet"
            archivo_resp_grupo = "onedrive:responsable_grupo_compras"
            archivo_centro = "onedrive:centro_sociedad_mro"
            archivo_mrp = "onedrive:responsable_mrp"

            if st.button("🔄 Forzar recarga desde OneDrive ahora"):
                loaders._descargar_onedrive.clear()
                st.session_state.pop("_clave_pipeline", None)
                st.rerun()

            if "onedrive" not in st.secrets:
                st.error(
                    "Falta configurar los Secrets de OneDrive (Settings → Secrets). "
                    "Mientras tanto, usa 'Subir archivos'."
                )
                st.stop()

        elif modo == "Subir archivos":
            archivo_data = st.file_uploader(
                "ME5A_con_Ariba (.xlsx o .parquet)",
                type=["xlsx", "parquet"],
                help="Si ya lo convertiste en la pestaña 'Preparar datos', sube el .parquet.",
            )
            archivo_resp_grupo = st.file_uploader("Responsable_Grupo_Compras.xlsx", type="xlsx")
            archivo_centro = st.file_uploader("Centro_Sociedad_MRO.xlsx", type="xlsx")
            archivo_mrp = st.file_uploader("Responsable_MRP.xlsx", type="xlsx")
            if not all([archivo_data, archivo_resp_grupo, archivo_centro, archivo_mrp]):
                st.info("Sube los 4 archivos para generar el reporte.")
                st.stop()
        else:
            archivo_data = "data/ME5A_con_Ariba.xlsx"
            archivo_resp_grupo = "data/Responsable_Grupo_Compras.xlsx"
            archivo_centro = "data/Centro_Sociedad_MRO.xlsx"
            archivo_mrp = "data/Responsable_MRP.xlsx"

        st.header("Parámetros")
        fecha_corte = st.date_input("Fecha de corte del reporte (FechaCorteReporte)", value=pd.Timestamp.today())
        st.caption(f"SLA: {config.SLA_DIAS_ERP_MRP} días ERP/MRP · {config.SLA_DIAS_ARIBA} días Ariba")

    def _clave_archivo(archivo):
        if hasattr(archivo, "name") and hasattr(archivo, "size"):
            return (archivo.name, archivo.size)
        return archivo

    clave_actual = (
        _clave_archivo(archivo_data),
        _clave_archivo(archivo_resp_grupo),
        _clave_archivo(archivo_centro),
        _clave_archivo(archivo_mrp),
        pd.Timestamp(fecha_corte),
        "df_trazabilidad_limpio" in st.session_state,
    )

    if st.session_state.get("_clave_pipeline") != clave_actual:
        df_data = loaders.cargar_data_pr(archivo_data)
        df_resp_grupo = loaders.cargar_responsable_grupo_compras(archivo_resp_grupo)
        df_centro_sociedad = loaders.cargar_centro_sociedad_mro(archivo_centro)
        df_resp_mrp = loaders.cargar_responsable_mrp(archivo_mrp)

        df_calculado = transform.pipeline_completo(
            df_data, df_resp_grupo, df_centro_sociedad, df_resp_mrp, fecha_corte=pd.Timestamp(fecha_corte)
        )
        df_calculado["Año"] = df_calculado["Fecha de pedido"].dt.year
        df_calculado["Mes"] = df_calculado["Fecha de pedido"].dt.month
        df_calculado["Día"] = df_calculado["Fecha de pedido"].dt.day

        if "df_trazabilidad_limpio" in st.session_state:
            df_traz = st.session_state["df_trazabilidad_limpio"]
            col_traz = "Solicitud de pedido" if "Solicitud de pedido" in df_traz.columns else "Solped SAP (600)"
            
            solpeds_no_cat = set(
                df_traz[col_traz]
                .dropna()
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
            )

            solpeds_me5a = (
                df_calculado["Solicitud de pedido"]
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
            )

            df_calculado["En_Trazabilidad"] = solpeds_me5a.isin(solpeds_no_cat)
        else:
            df_calculado["En_Trazabilidad"] = False

        df_calculado["Tipo Ariba"] = df_calculado.apply(determinar_tipo_ariba, axis=1)

        st.session_state["_df_pipeline"] = df_calculado
        st.session_state["_clave_pipeline"] = clave_actual

    df = st.session_state["_df_pipeline"]

    NS_MIN_GLOBAL = int(df["Nivel de Servicio"].min()) if len(df) and pd.notna(df["Nivel de Servicio"].min()) else 0
    NS_MAX_GLOBAL = int(df["Nivel de Servicio"].max()) if len(df) and pd.notna(df["Nivel de Servicio"].max()) else 100

    checkpoints = [("0. Total tras el pipeline (sin filtros)", len(df))]
    metricas_por_etapa = []

    def _snapshot(nombre, d):
        n = len(d)
        pct = (d["Cumple"] == "Cumple").sum() / n * 100 if n else 0
        prom = d["Nivel de Servicio"].mean() if n else float("nan")
        pos_oc = d["Pedido"].nunique() + (1 if d["Pedido"].isna().any() else 0)
        metricas_por_etapa.append(
            (nombre, f"{n:,}", f"{pct:.0f}%", f"{prom:.0f}" if pd.notna(prom) else "-", f"{pos_oc:,}")
        )

    _snapshot("0. Sin filtros", df)

    # ---- Filtros ----
    st.subheader("Filtros")
    c1, c2, c3 = st.columns(3)
    with c1:
        centros = st.multiselect("Centro", sorted(df["Centro"].dropna().unique()))
    with c2:
        aplica = st.multiselect("Aplica?", sorted(df["Aplica?"].dropna().unique()))
    with c3:
        tipos_ariba = st.multiselect("Origen / Tipo Solicitud", sorted(df["Tipo Ariba"].dropna().unique()))

    df_f = df.copy()
    if centros:
        df_f = df_f[df_f["Centro"].isin(centros)]
    checkpoints.append(("1. Tras filtro Centro", len(df_f)))

    if aplica:
        df_f = df_f[df_f["Aplica?"].isin(aplica)]
    checkpoints.append(("2. Tras filtro Aplica?", len(df_f)))

    if tipos_ariba:
        df_f = df_f[df_f["Tipo Ariba"].isin(tipos_ariba)]
    checkpoints.append(("2b. Tras filtro Origen / Tipo Solicitud", len(df_f)))

    _snapshot("2. Tras Centro + Aplica? + Origen", df_f)

    # ---- Estado Solped ----
    st.caption("Estado Solped (el filtro de fecha de abajo solo aplica dentro de 'Pedido completo')")
    h1, h2, h3, h4 = st.columns(4)
    with h1:
        estados = st.multiselect("Estado Solped", sorted(df_f["Estado Solped"].dropna().unique()))

    if estados:
        df_f = df_f[df_f["Estado Solped"].isin(estados)]
    checkpoints.append(("3. Tras filtro Estado Solped (total)", len(df_f)))
    checkpoints.append(("3a. Sin pedido (antes de jerarquía)", (df_f["Estado Solped"] == "Sin pedido").sum()))
    checkpoints.append(("3b. Pedido incompleto (antes de jerarquía)", (df_f["Estado Solped"] == "Pedido incompleto").sum()))
    checkpoints.append(("3c. Pedido completo (antes de jerarquía)", (df_f["Estado Solped"] == "Pedido completo").sum()))
    _snapshot("3. Tras Estado Solped", df_f)

    # ---- Jerarquía Año / Mes / Día ----
    df_pedido_completo = df_f[df_f["Estado Solped"] == "Pedido completo"]

    with h2:
        años = st.multiselect("Año", sorted(df_pedido_completo["Año"].dropna().unique().astype(int)))
    with h3:
        _base_mes = df_pedido_completo[df_pedido_completo["Año"].isin(años)] if años else df_pedido_completo
        meses = st.multiselect("Mes", sorted(_base_mes["Mes"].dropna().unique().astype(int)))
    with h4:
        _base_dia = _base_mes[_base_mes["Mes"].isin(meses)] if meses else _base_mes
        fechas_disponibles = sorted(_base_dia["Fecha de pedido"].dt.date.dropna().unique())
        fechas = st.multiselect("Día", fechas_disponibles, format_func=lambda f: f.strftime("%d-%m-%Y"))

    if años or meses or fechas:
        es_pedido_completo = df_f["Estado Solped"] == "Pedido completo"
        cond_fecha = pd.Series(True, index=df_f.index)
        if años:
            cond_fecha &= df_f["Año"].isin(años)
        if meses:
            cond_fecha &= df_f["Mes"].isin(meses)
        if fechas:
            cond_fecha &= df_f["Fecha de pedido"].dt.date.isin(fechas)
        df_f = df_f[~es_pedido_completo | (es_pedido_completo & cond_fecha)]

    checkpoints.append(("4a. Sin pedido tras jerarquía", (df_f["Estado Solped"] == "Sin pedido").sum()))
    checkpoints.append(("4b. Pedido incompleto tras jerarquía", (df_f["Estado Solped"] == "Pedido incompleto").sum()))
    checkpoints.append(("4c. Pedido completo tras jerarquía", (df_f["Estado Solped"] == "Pedido completo").sum()))
    checkpoints.append(("4. Total tras jerarquía de fecha", len(df_f)))
    _snapshot("4. Tras jerarquía Año/Mes/Día", df_f)

    # ---- Solped MRP y Cumple ----
    c4, c5 = st.columns(2)
    with c4:
        solped_mrp = st.multiselect("Solped MRP", sorted(df_f["Solped MRP"].dropna().unique()))
    with c5:
        cumple = st.multiselect("Nivel de Servicio (Cumple)", sorted(df_f["Cumple"].dropna().unique()))

    if solped_mrp:
        df_f = df_f[df_f["Solped MRP"].isin(solped_mrp)]
    checkpoints.append(("5. Tras filtro Solped MRP", len(df_f)))
    _snapshot("5. Tras Solped MRP", df_f)

    if cumple:
        df_f = df_f[df_f["Cumple"].isin(cumple)]
    checkpoints.append(("6. Tras filtro Cumple", len(df_f)))
    _snapshot("6. Tras Cumple", df_f)

    st.divider()
    st.subheader("Filtro por días de gestión (Nivel de Servicio)")

    if len(df_f):
        n_negativos = (df_f["Nivel de Servicio"] < 0).sum()

        f1, f2 = st.columns([1, 2])
        with f1:
            excluir_negativos = st.checkbox(
                "Excluir negativos (solo desde 0)",
                value=False,
                help="Filas con días negativos por modificación posterior a la OC.",
            )
        with f2:
            if excluir_negativos:
                rango_ns = (0, NS_MAX_GLOBAL)
                st.caption(f"Rango aplicado: 0 a {NS_MAX_GLOBAL:,} días (negativos excluidos)")
            elif NS_MIN_GLOBAL < NS_MAX_GLOBAL:
                rango_ns = st.slider(
                    "Rango de días de gestión",
                    min_value=NS_MIN_GLOBAL,
                    max_value=NS_MAX_GLOBAL,
                    value=(NS_MIN_GLOBAL, NS_MAX_GLOBAL),
                    key="slider_rango_dias",
                )
            else:
                rango_ns = (NS_MIN_GLOBAL, NS_MAX_GLOBAL)
                st.caption(f"Todas las filas tienen {NS_MIN_GLOBAL} días")

        df_f = df_f[df_f["Nivel de Servicio"].between(rango_ns[0], rango_ns[1])]

    checkpoints.append(("7. Tras filtro Nivel de Servicio (RESULTADO FINAL)", len(df_f)))
    _snapshot("7. RESULTADO FINAL", df_f)

    # ---- Diagnóstico de filtrado ----
    with st.expander("🔍 Diagnóstico de filtrado (para comparar contra el pbix)"):
        st.write("Filas en cada etapa:")
        st.table(pd.DataFrame(checkpoints, columns=["Etapa", "Filas"]))
        st.write("**Las 3 métricas en cada etapa del filtro**:")
        st.table(pd.DataFrame(metricas_por_etapa, columns=["Etapa", "Filas", "% Cumplimiento", "Prom. días", "Pos. OC"]))
        st.dataframe(
            df_f[["Solicitud de pedido", "Centro", "Estado Solped", "Fecha de pedido", "Nivel de Servicio", "Cumple", "Solped MRP", "Tipo Ariba"]]
            .sort_values("Solicitud de pedido"),
            use_container_width=True,
        )
        csv = df_f.to_csv(index=False).encode("utf-8")
        st.download_button("Descargar detalle filtrado (CSV)", csv, "detalle_filtrado.csv", "text/csv")

    # ---- Tarjetas KPI ----
    VERDE, VERDE_BORDE = "rgba(35, 145, 75, 0.16)", "rgba(35, 145, 75, 0.55)"
    ROJO, ROJO_BORDE = "rgba(204, 0, 0, 0.14)", "rgba(204, 0, 0, 0.55)"

    pct_cumplimiento = (df_f["Cumple"] == "Cumple").sum() / max(len(df_f), 1) * 100
    promedio_dias = df_f["Nivel de Servicio"].mean()
    promedio_lead_time = df_f["Lead Time Total"].mean() if "Lead Time Total" in df_f.columns else float("nan")
    pedidos_distintos = df_f["Pedido"].nunique() + (1 if df_f["Pedido"].isna().any() else 0)

    # Condición de texto para las Cards
    color_texto_card = "#FF3333" if st.session_state["tema"] == "oscuro" else "#404B55"
    color_sub_card = "#FF3333" if st.session_state["tema"] == "oscuro" else "#555"

    def tarjeta(titulo: str, valor: str, subtitulo: str = "", fondo: str = "rgba(64,75,85,0.07)", borde: str = "rgba(64,75,85,0.35)") -> str:
        html_sub = f'<div style="font-size:0.78rem;color:{color_sub_card};margin-top:4px;font-weight:500;">{subtitulo}</div>' if subtitulo else ""
        return (
            f'<div style="background:{fondo};border:1.5px solid {borde};border-radius:8px;'
            f'padding:12px 18px;text-align:center;">'
            f'<div style="font-size:0.78rem;color:{color_texto_card};font-weight:600;letter-spacing:.03em;'
            f'text-transform:uppercase;opacity:.85;margin-bottom:4px;">{titulo}</div>'
            f'<div style="font-size:2rem;font-weight:700;color:{color_texto_card};line-height:1.1;">{valor}</div>'
            f'{html_sub}'
            f'</div>'
        )

    if pd.isna(promedio_dias):
        f_dias, b_dias, txt_dias = "rgba(64,75,85,0.07)", "rgba(64,75,85,0.35)", "-"
    elif promedio_dias > 10:
        f_dias, b_dias, txt_dias = ROJO, ROJO_BORDE, f"{promedio_dias:.0f}"
    else:
        f_dias, b_dias, txt_dias = VERDE, VERDE_BORDE, f"{promedio_dias:.0f}"

    f_pct, b_pct = (VERDE, VERDE_BORDE) if pct_cumplimiento >= 85 else (ROJO, ROJO_BORDE)
    txt_lt = f"{promedio_lead_time:.0f}" if pd.notna(promedio_lead_time) else "-"

    t1, t2, t3 = st.columns(3)
    with t1:
        st.markdown(
            tarjeta(
                "Nivel de Servicio",
                f"{txt_dias} días",
                subtitulo=f"Lead Time Total: <b>{txt_lt}</b> días",
                fondo=f_dias,
                borde=b_dias,
            ),
            unsafe_allow_html=True,
        )
    with t2:
        st.markdown(tarjeta("% Cumplimiento SLA", f"{pct_cumplimiento:.0f}%", fondo=f_pct, borde=b_pct), unsafe_allow_html=True)
    with t3:
        st.markdown(tarjeta("OC generadas", f"{pedidos_distintos:,}"), unsafe_allow_html=True)

    st.divider()

    # ---- Estilo corporativo Enaex ----
    ENAEX_GRIS = "#1E2329" if st.session_state["tema"] == "oscuro" else "#404B55"
    ENAEX_ROJO = "#CC0000"

    def tabla_enaex(tabla: pd.DataFrame, max_height: int | None = None, compacta: bool = False) -> str:
        cols = list(tabla.columns)

        def _fmt(col, val):
            if pd.isna(val):
                return "-"
            if col in ["Promedio días de gestión", "Promedio Lead Time Total"]:
                return f"{val:,.0f}"
            if col == "% Cumplimiento":
                return f"{val:,.0f}%"
            if col == "Pos. OC generadas":
                return f"{val:,.0f}"
            return str(val)

        pad = "4px 6px" if compacta else "7px 12px"
        pad_th = "5px 6px" if compacta else "9px 12px"
        fuente = "0.72rem" if compacta else "0.86rem"
        fuente_th = "0.66rem" if compacta else "0.82rem"

        color_texto_tabla = "#FF3333" if st.session_state["tema"] == "oscuro" else "#404B55"

        filas = []
        for i, r in enumerate(tabla.itertuples(index=False)):
            es_total = str(r[0]) == "TOTAL"
            if es_total:
                estilo_fila = f"background:{ENAEX_GRIS};color:#fff;font-weight:700;border-top:2px solid {ENAEX_ROJO};"
            else:
                fondo = "#ffffff" if (i % 2 == 0 and st.session_state["tema"] == "claro") else ("#f4f5f7" if st.session_state["tema"] == "claro" else "#14181d")
                estilo_fila = f"background:{fondo};color:{color_texto_tabla};"
            celdas = []
            for j, c in enumerate(cols):
                align = "left" if j == 0 else "right"
                celdas.append(
                    f'<td style="padding:{pad};text-align:{align};'
                    f'border-bottom:1px solid #d8dbdf;white-space:nowrap;">{_fmt(c, r[j])}</td>'
                )
            filas.append(f'<tr style="{estilo_fila}">{"".join(celdas)}</tr>')

        abrev = {
            "Promedio días de gestión": "Nivel Serv.",
            "Promedio Lead Time Total": "LT Total",
            "% Cumplimiento": "% Cumpl.",
            "Pos. OC generadas": "Pos. OC",
        }
        encabezados = "".join(
            f'<th style="padding:{pad_th};text-align:{"left" if j == 0 else "right"};'
            f'background:{ENAEX_GRIS};color:#fff;font-weight:600;font-size:{fuente_th};'
            f'letter-spacing:.02em;position:sticky;top:0;z-index:2;white-space:nowrap;">'
            f"{abrev.get(c, c) if compacta else c}</th>"
            for j, c in enumerate(cols)
        )

        cuerpo_tabla = "".join(filas)
        ancho_min = "340px" if compacta else "auto"

        tabla_html = (
            f'<table style="width:100%;min-width:{ancho_min};border-collapse:collapse;font-size:{fuente};'
            f'font-family:inherit;border:1px solid #d8dbdf;">'
            f'<thead><tr>{encabezados}</tr></thead>'
            f'<tbody>{cuerpo_tabla}</tbody>'
            f'</table>'
        )

        alto = f"max-height:{max_height}px;overflow-y:auto;" if max_height else ""
        return f'<div style="{alto}overflow-x:auto;border:1px solid #d8dbdf;border-radius:4px;">{tabla_html}</div>'

    # ---- Tabla por Comprador ----
    st.subheader("Por comprador")
    st.caption(
        "Asignación por **grupo de compras**: las líneas MRP se reparten entre los "
        "compradores responsables de cada grupo, en vez de concentrarse en el responsable de MRP."
    )
    col_comprador = (
        "Comprador (Grupo de compras)"
        if "Comprador (Grupo de compras)" in df_f.columns
        else "Comprador por Grupo Compras"
    )
    tabla_comprador = transform.calcular_metricas_por_grupo(df_f, [col_comprador])
    tabla_comprador = transform.agregar_fila_total(tabla_comprador, df_f, [col_comprador])
    st.markdown(tabla_enaex(tabla_comprador), unsafe_allow_html=True)

    st.write("")

    # ---- Dos vistas por centro en paralelo ----
    vc1, vc2 = st.columns(2)

    with vc1:
        st.subheader("Por centro logístico")
        st.caption("Vista fija — el total calza con la vista por comprador.")
        tabla_fija = transform.tabla_centros_fija(df_f)
        st.markdown(tabla_enaex(tabla_fija, compacta=True), unsafe_allow_html=True)

    with vc2:
        st.subheader("Detalle por centro")
        st.caption("Centros activos según los filtros aplicados.")
        cols_detalle = [c for c in ["Centro", "Nombre Centro 2"] if c in df_f.columns]
        tabla_detalle = transform.calcular_metricas_por_grupo(df_f, cols_detalle)
        tabla_detalle = tabla_detalle.sort_values("Pos. OC generadas", ascending=False)
        tabla_detalle = transform.agregar_fila_total(tabla_detalle, df_f, cols_detalle)
        st.markdown(tabla_enaex(tabla_detalle, max_height=300, compacta=True), unsafe_allow_html=True)

    st.divider()

    # ---- Detalle de solicitudes ----
    COLUMNAS_DETALLE = [
        "Centro",
        "Material",
        "Texto breve",
        "Solicitud de pedido",
        "Tipo Ariba",
        "Fecha de solicitud",
        "Fecha de liberación",
        "Fecha modificación",
        "Grupo de compras",
        "Cantidad pedida",
        "Pedido",
        "Fecha de pedido",
        "Posición de pedido",
        "Comprador (Grupo de compras)",
        "Solped MRP",
        "Nombre Centro 2",
        "Nombre Centro",
        "Estado Solped",
        "Nivel de Servicio",
        "Lead Time Total",
        "Comentario",
        "Cumple",
    ]

    def preparar_detalle(d: pd.DataFrame) -> pd.DataFrame:
        d = d.copy()
        for col_fecha in ["Fecha de solicitud", "Fecha de liberación", "Fecha modificación", "Fecha de pedido"]:
            if col_fecha in d.columns:
                d[col_fecha] = pd.to_datetime(d[col_fecha], errors="coerce").dt.date
        if "Comentario" not in d.columns:
            d["Comentario"] = ""
        return d[[c for c in COLUMNAS_DETALLE if c in d.columns]]

    detalle = preparar_detalle(df_f)

    with st.expander("Ver detalle de solicitudes", expanded=False):
        st.caption("La columna **Comentario** es editable: escribe ahí y se incluirá en el Excel de registro.")
        detalle_editado = st.data_editor(
            detalle,
            use_container_width=True,
            num_rows="fixed",
            key="editor_detalle",
            column_config={
                "Tipo Ariba": st.column_config.TextColumn("Origen / Tipo Solicitud", width="medium"),
                "Comentario": st.column_config.TextColumn(
                    "Comentario", help="Anotación libre para el registro semanal", width="medium"
                ),
                "Nivel de Servicio": st.column_config.NumberColumn("Nivel de Servicio (días)"),
                "Lead Time Total": None,  # 👈 Oculto visualmente del frontend
            },
            disabled=[c for c in detalle.columns if c != "Comentario"],
        )

    # ---- Exportación a Excel ----
    semana_ref = pd.Timestamp(fecha_corte) - pd.Timedelta(days=7)
    num_semana = semana_ref.isocalendar()[1]
    nombre_archivo = f"Sem{num_semana:02d}-{semana_ref.year}.xlsx"

    def generar_excel(detalle_df: pd.DataFrame) -> bytes:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        gris = "FF404B55"
        rojo = "FFCC0000"
        fuente_base = "Arial"

        wb = Workbook()
        borde = Border(bottom=Side(style="thin", color="FFD8DBDF"))

        def escribir_hoja(ws, titulo, tabla, col_inicio=1, fila_inicio=1):
            ws.cell(row=fila_inicio, column=col_inicio, value=titulo).font = Font(
                name=fuente_base, bold=True, size=12, color=gris
            )
            fila = fila_inicio + 1
            for j, col in enumerate(tabla.columns):
                c = ws.cell(row=fila, column=col_inicio + j, value=str(col))
                c.font = Font(name=fuente_base, bold=True, color="FFFFFFFF", size=10)
                c.fill = PatternFill("solid", fgColor=gris)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for i, r in enumerate(tabla.itertuples(index=False, name=None)):
                es_total = str(r[0]) == "TOTAL"
                for j, col in enumerate(tabla.columns):
                    val = r[j]
                    if pd.isna(val):
                        val = None
                    elif isinstance(val, (int, float)) and col in (
                        "Promedio días de gestión",
                        "Promedio Lead Time Total",
                        "% Cumplimiento",
                        "Pos. OC generadas",
                    ):
                        val = round(float(val))
                    elif hasattr(val, "item"):
                        val = val.item()
                    c = ws.cell(row=fila + 1 + i, column=col_inicio + j, value=val)
                    c.font = Font(name=fuente_base, size=10, bold=es_total, color=gris)
                    c.border = borde
                    if es_total:
                        c.fill = PatternFill("solid", fgColor="FFEFF0F2")
                    if col == "% Cumplimiento" and val is not None:
                        c.number_format = '0"%"'
                    if col in ("Fecha de solicitud", "Fecha de liberación", "Fecha modificación", "Fecha de pedido"):
                        c.number_format = "DD-MM-YYYY"
            return fila + 1 + len(tabla)

        def ajustar_ancho(ws, tabla, col_inicio=1, extra=3):
            for j, col in enumerate(tabla.columns):
                largos = [len(str(col))]
                for v in tabla[col].head(200):
                    largos.append(0 if pd.isna(v) else len(str(v)))
                ws.column_dimensions[get_column_letter(col_inicio + j)].width = min(max(largos) + extra, 45)

        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = f"Nivel de Servicio MRO — Registro semana {num_semana:02d}/{semana_ref.year}"
        ws["A1"].font = Font(name=fuente_base, bold=True, size=14, color=gris)
        ws["A2"] = f"Fecha de corte del reporte: {pd.Timestamp(fecha_corte).date()}"
        ws["A2"].font = Font(name=fuente_base, size=10, italic=True, color=gris)
        ws["A3"] = f"Generado: {pd.Timestamp.today().date()}"
        ws["A3"].font = Font(name=fuente_base, size=10, italic=True, color=gris)

        promedio_lt_val = round(promedio_lead_time) if pd.notna(promedio_lead_time) else None
        resumen = pd.DataFrame(
            {
                "Indicador": [
                    "Promedio Nivel de Servicio (días)",
                    "Promedio Lead Time Total (días)",
                    "% Cumplimiento SLA",
                    "OC generadas",
                    "Líneas consideradas",
                ],
                "Valor": [
                    round(promedio_dias) if pd.notna(promedio_dias) else None,
                    promedio_lt_val,
                    round(pct_cumplimiento),
                    pedidos_distintos,
                    len(df_f),
                ],
            }
        )
        fila = escribir_hoja(ws, "Indicadores generales", resumen, fila_inicio=5)
        ws.cell(row=5, column=1).font = Font(name=fuente_base, bold=True, size=12, color=rojo)
        ajustar_ancho(ws, resumen)

        fila = escribir_hoja(ws, "Por comprador", tabla_comprador, fila_inicio=fila + 2)
        fila = escribir_hoja(ws, "Por centro logístico", tabla_fija, fila_inicio=fila + 2)
        escribir_hoja(ws, "Detalle por centro", tabla_detalle, fila_inicio=fila + 2)

        ws2 = wb.create_sheet("Detalle solicitudes")
        escribir_hoja(ws2, "Detalle de solicitudes", detalle_df)
        ajustar_ancho(ws2, detalle_df)
        ws2.freeze_panes = "A3"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    st.subheader("Registro semanal")
    st.caption(f"Prepara el reporte completo (indicadores, tablas y detalle con comentarios) como **{nombre_archivo}**.")

    clave_excel = (len(df_f), int(pd.util.hash_pandas_object(detalle_editado["Comentario"].fillna("")).sum()), pd.Timestamp(fecha_corte))

    col_prep, col_desc = st.columns([1, 2])
    with col_prep:
        if st.button("📄 Preparar Excel"):
            with st.spinner("Generando el Excel..."):
                try:
                    st.session_state["_excel_bytes"] = generar_excel(detalle_editado)
                    st.session_state["_excel_clave"] = clave_excel
                except Exception as e:
                    st.error(f"No se pudo generar el Excel: {e}")

    with col_desc:
        excel_listo = st.session_state.get("_excel_bytes") is not None
        excel_desactualizado = st.session_state.get("_excel_clave") != clave_excel
        if excel_listo and excel_desactualizado:
            st.caption("⚠️ Los filtros cambiaron desde que preparaste este Excel — vuelve a preparar para reflejar la selección actual.")
        if excel_listo:
            st.download_button(
                f"⬇ Descargar {nombre_archivo}",
                data=st.session_state["_excel_bytes"],
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
            )
        else:
            st.caption('Haz clic en "Preparar Excel" para generar el archivo de descarga.')


# ==============================================================================
# PESTAÑA 2: TRAZABILIDAD NO CATALOGADAS
# ==============================================================================
with tab_trazabilidad:
    st.title("🔍 Trazabilidad PR No Catalogadas — Ariba")

    if not HAS_TRAZABILIDAD:
        st.warning(
            "⚠️ **Módulo 'ariba_trazabilidad.py' no disponible.**\n\n"
            "Asegúrate de que el archivo esté subido en GitHub en la misma carpeta raíz."
        )
    else:
        st.markdown(
            "Procesa el reporte de **Trazabilidad Ariba** sin consolidar para vincular "
            "las solicitudes de compra iniciales, sus agregadas y su salida a SAP ERP (Solped 600)."
        )

        traz_col1, traz_col2 = st.columns([2, 1])
        with traz_col1:
            archivo_trazabilidad = st.file_uploader(
                "Cargar Reporte PR No Catalogadas - Trazabilidad (.csv)",
                type=["csv"],
                key="uploader_trazabilidad",
            )
        with traz_col2:
            empresa_id = st.text_input(
                "Empresa compradora (ID)",
                value=getattr(trazabilidad, "EMPRESA_POR_DEFECTO", "1000"),
            )

        if archivo_trazabilidad:
            if st.button("🚀 Procesar Trazabilidad", key="btn_procesar_traz"):
                with st.spinner("Procesando trazabilidad y reconstruyendo cadena de eventos..."):
                    try:
                        df_cadena, df_resumen = trazabilidad.procesar_trazabilidad_completa(
                            archivo_trazabilidad, empresa=empresa_id
                        )

                        st.session_state["df_trazabilidad_cadena"] = df_cadena
                        st.session_state["df_trazabilidad_limpio"] = df_resumen

                        # Borrar la caché del pipeline para obligar a recalcular la Pestaña 1
                        st.session_state.pop("_clave_pipeline", None)
                        st.success("¡Trazabilidad procesada con éxito! La pestaña Dx Compradores ha sido actualizada.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al procesar el archivo: {e}")

        if "df_trazabilidad_limpio" in st.session_state and "df_trazabilidad_cadena" in st.session_state:
            df_cadena = st.session_state["df_trazabilidad_cadena"]
            df_resumen = st.session_state["df_trazabilidad_limpio"]

            kpi_t1, kpi_t2, kpi_t3 = st.columns(3)
            with kpi_t1:
                st.metric("Total PR Iniciales", f"{len(df_cadena):,}")
            with kpi_t2:
                completas = (df_cadena["Cadena completa"] == "Sí").sum() if "Cadena completa" in df_cadena.columns else 0
                st.metric("Cadenas Completas (con SAP)", f"{completas:,}")
            with kpi_t3:
                st.metric("Solpeds 600 Identificadas", f"{len(df_resumen):,}")

            st.divider()

            tab_cad, tab_res = st.tabs(["📋 Cadena Detallada", "📊 Resumen por Solped (SAP 600)"])

            with tab_cad:
                st.subheader("Vista Cadena de Eventos")
                st.dataframe(df_cadena, use_container_width=True)
                csv_cadena = df_cadena.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇ Descargar Cadena Completa (CSV)",
                    data=csv_cadena,
                    file_name="Trazabilidad_Cadena_Completa.csv",
                    mime="text/csv",
                )

            with tab_res:
                st.subheader("Vista Consolidada por Solped SAP 600")
                st.caption("Esta información servirá de cruce directo con la base ME5A.")
                st.dataframe(df_resumen, use_container_width=True)
                csv_resumen = df_resumen.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇ Descargar Resumen Solpeds 600 (CSV)",
                    data=csv_resumen,
                    file_name="Resumen_Solped_600_No_Catalogadas.csv",
                    mime="text/csv",
                )
