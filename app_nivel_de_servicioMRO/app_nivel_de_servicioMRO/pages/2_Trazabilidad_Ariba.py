"""
Página: Trazabilidad de PR No Catalogadas (Ariba)

Toma el reporte 'PR No Catalogadas - Trazabilidad' exportado de Ariba y
reconstruye la cadena completa:

    PR inicial  ->  toma del comprador  ->  PR agregada  ->  Solped SAP (600) + PO

Esta página es independiente del dashboard de Nivel de Servicio: todavía no
cruza con la ME5A. Sirve para dejar el reporte de Ariba en un formato limpio
y medir los tramos de gestión que SAP no registra.
"""
import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st

# Permite importar los módulos que están un nivel más arriba (junto a app.py)
sys.path.append(str(Path(__file__).resolve().parent.parent))
import ariba_trazabilidad as at  # noqa: E402

st.set_page_config(page_title="Trazabilidad PR Ariba", layout="wide")

ENAEX_GRIS = "#404B55"
ENAEX_ROJO = "#CC0000"

st.title("Trazabilidad de PR No Catalogadas (Ariba)")
st.caption(
    "Reconstruye la cadena PR inicial → toma del comprador → PR agregada → Solped SAP + PO, "
    "para medir el tiempo de gestión que no queda registrado en SAP."
)

# ---------------- Carga ----------------
with st.sidebar:
    st.header("Archivo de entrada")
    archivo = st.file_uploader("Reporte PR No Catalogadas - Trazabilidad (.csv)", type="csv")
    if not archivo:
        st.info("Sube el reporte exportado de Ariba para continuar.")

if not archivo:
    st.stop()

try:
    empresas = at.empresas_disponibles(archivo)
except Exception as e:
    st.error(f"No se pudo leer el archivo: {e}")
    st.stop()

with st.sidebar:
    empresa = st.selectbox(
        "Empresa compradora",
        empresas,
        index=empresas.index(at.EMPRESA_POR_DEFECTO) if at.EMPRESA_POR_DEFECTO in empresas else 0,
    )

archivo.seek(0)
datos = at.cargar_reporte(archivo, empresa=empresa)
cadena = at.construir_cadena(datos)

# ---------------- Filtros ----------------
st.subheader("Filtros")
f1, f2, f3 = st.columns(3)
with f1:
    solo_completas = st.selectbox(
        "Cadena", ["Todas", "Solo cadena completa (con Solped SAP)", "Solo sin cadena"], index=0
    )
with f2:
    estados = st.multiselect(
        "Estado de agregación", sorted(cadena["Estado de agregación"].dropna().unique())
    )
with f3:
    fechas_validas = cadena["Fecha creación PR inicial"].dropna()
    if len(fechas_validas):
        rango = st.date_input(
            "Rango de creación de la PR inicial",
            value=(fechas_validas.min().date(), fechas_validas.max().date()),
        )
    else:
        rango = None

c = cadena.copy()
if solo_completas.startswith("Solo cadena completa"):
    c = c[c["Cadena completa"] == "Sí"]
elif solo_completas.startswith("Solo sin"):
    c = c[c["Cadena completa"] == "No"]
if estados:
    c = c[c["Estado de agregación"].isin(estados)]
if rango and isinstance(rango, (tuple, list)) and len(rango) == 2:
    c = c[
        c["Fecha creación PR inicial"].between(
            pd.Timestamp(rango[0]), pd.Timestamp(rango[1]) + pd.Timedelta(days=1)
        )
    ]

# ---------------- Indicadores ----------------
def tarjeta(titulo: str, valor: str, detalle: str = "") -> str:
    extra = (
        f'<div style="font-size:0.72rem;color:{ENAEX_GRIS};opacity:.7;margin-top:2px;">{detalle}</div>'
        if detalle
        else ""
    )
    return (
        f'<div style="background:rgba(64,75,85,0.07);border:1.5px solid rgba(64,75,85,0.35);'
        f'border-radius:8px;padding:12px 16px;text-align:center;">'
        f'<div style="font-size:0.74rem;color:{ENAEX_GRIS};font-weight:600;letter-spacing:.03em;'
        f'text-transform:uppercase;opacity:.85;margin-bottom:4px;">{titulo}</div>'
        f'<div style="font-size:1.7rem;font-weight:700;color:{ENAEX_GRIS};line-height:1.1;">{valor}</div>'
        f"{extra}</div>"
    )


completas = c[c["Cadena completa"] == "Sí"]


def _mediana(col):
    v = completas[col].dropna() if len(completas) else pd.Series(dtype=float)
    return f"{v.median():.1f}" if len(v) else "-"


k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(tarjeta("PR iniciales", f"{len(c):,}"), unsafe_allow_html=True)
with k2:
    st.markdown(
        tarjeta("Con cadena completa", f"{len(completas):,}", "llegan a Solped SAP + PO"),
        unsafe_allow_html=True,
    )
with k3:
    st.markdown(
        tarjeta("Días espera en panel", _mediana("Días espera en panel"), "mediana"),
        unsafe_allow_html=True,
    )
with k4:
    st.markdown(
        tarjeta("Días gestión total", _mediana("Días gestión total"), "mediana · liberación → PO"),
        unsafe_allow_html=True,
    )

st.caption(
    "**Espera en panel**: desde que se libera la PR inicial hasta que el comprador la toma. "
    "**Gestión total**: desde la liberación de la PR inicial hasta la orden de compra."
)

st.divider()

# ---------------- Cadena ----------------
st.subheader("Cadena de trazabilidad")
st.caption("Se lee de izquierda a derecha: PR inicial → toma del comprador → PR agregada → salida a SAP.")

formato_fechas = {
    col: st.column_config.DatetimeColumn(col, format="DD-MM-YYYY HH:mm")
    for col in c.columns
    if col.startswith("Fecha") and "asignada" in col.lower()
}
formato_fechas.update(
    {
        col: st.column_config.DateColumn(col, format="DD-MM-YYYY")
        for col in c.columns
        if col.startswith("Fecha") and "asignada" not in col.lower()
    }
)

st.dataframe(c, use_container_width=True, height=420, column_config=formato_fechas)

# ---------------- Resumen por Solped ----------------
resumen = at.resumen_por_solped(c)
with st.expander(f"Resumen por Solped SAP ({len(resumen):,} solpeds) — base para el cruce futuro con la ME5A"):
    st.caption(
        "Una fila por código 600. Cuando varias PR iniciales se consolidan en una agregada, "
        "se toma la liberación más antigua: es cuando arrancó el reloj del comprador."
    )
    st.dataframe(resumen, use_container_width=True, height=320)

# ---------------- Descarga ----------------
st.divider()
st.subheader("Descargar")


def generar_excel(cadena_df: pd.DataFrame, resumen_df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        cadena_df.to_excel(writer, sheet_name="Cadena trazabilidad", index=False)
        resumen_df.to_excel(writer, sheet_name="Resumen por Solped", index=False)
        wb = writer.book
        for hoja, tabla in [
            ("Cadena trazabilidad", cadena_df),
            ("Resumen por Solped", resumen_df),
        ]:
            ws = wb[hoja]
            for j, col in enumerate(tabla.columns, start=1):
                celda = ws.cell(row=1, column=j)
                celda.font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
                celda.fill = PatternFill("solid", fgColor="FF404B55")
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                largos = [len(str(col))]
                for v in tabla[col].head(200):
                    largos.append(0 if pd.isna(v) else len(str(v)))
                ws.column_dimensions[get_column_letter(j)].width = min(max(largos) + 3, 32)
            ws.freeze_panes = "A2"
    return buffer.getvalue()


try:
    st.download_button(
        "⬇ Descargar cadena de trazabilidad (Excel)",
        data=generar_excel(c, resumen),
        file_name=f"Trazabilidad_Ariba_{empresa}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.error(f"No se pudo generar el Excel: {e}")
