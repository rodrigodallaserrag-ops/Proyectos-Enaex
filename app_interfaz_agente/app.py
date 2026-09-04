import streamlit as st
import pandas as pd
import numpy as np
import re
import requests
from datetime import datetime, date
import io

# Librerías para diseño de Excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Librería para generación de PDF
from fpdf import FPDF

# =============================================================================
# CONFIGURACIÓN DE PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Sistema Integrado de Evaluación de Ofertas - Enaex",
    page_icon="⚡",
    layout="wide"
)

# Estilos CSS personalizados
st.markdown("""
<style>
    .main-header { font-size: 1.8rem; font-weight: 700; color: #1E3A8A; margin-bottom: 0.5rem; }
    .sub-header { font-size: 1rem; color: #4B5563; margin-bottom: 1.5rem; }
    .stTable { font-size: 0.85rem; }
    .metric-card { background-color: #F3F4F6; padding: 1rem; border-radius: 0.5rem; border-left: 4px solid #1E3A8A; }
    .badge-best { background-color: #D1FAE5; color: #065F46; padding: 0.2rem 0.5rem; border-radius: 0.25rem; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# OBTENCIÓN DE INDICADORES FINANCIEROS EN TIEMPO REAL (API)
# =============================================================================
@st.cache_data(ttl=3600)
def obtener_indicadores_tiempo_real():
    """Consulta la API de mindicador.cl para obtener USD, EUR y UF actualizados"""
    valores_defecto = {"USD": 950.0, "EUR": 1020.0, "UF": 38000.0, "estado": False}
    try:
        response = requests.get("https://mindicador.cl/api", timeout=5)
        if response.status_code == 200:
            data = response.json()
            return {
                "USD": float(data.get("dolar", {}).get("valor", 950.0)),
                "EUR": float(data.get("euro", {}).get("valor", 1020.0)),
                "UF": float(data.get("uf", {}).get("valor", 38000.0)),
                "estado": True
            }
    except Exception:
        pass
    return valores_defecto

# =============================================================================
# FUNCIONES DE EXPORTACIÓN Y FORMATO (EXCEL Y PDF)
# =============================================================================
def generar_excel_estilizado(df, moneda_vista):
    """Genera un archivo Excel con diseño corporativo elegante, bordes, formatos numéricos y anchos ajustados."""
    buffer_excel = io.BytesIO()
    
    cols_export = [
        'SOLPED', 'Pos', 'Material', 'Centro', 'Cantidad', 'UM', 
        'Precio Unitario', 'Moneda', 'Proveedor Visual', 
        'Calendario de entrega', 'Días para Entrega', 'Monto Total Visualizado'
    ]
    
    df_export = df[[c for c in cols_export if c in df.columns]].copy()
    df_export.rename(columns={
        'Proveedor Visual': 'Proveedor', 
        'Monto Total Visualizado': f'Total ({moneda_vista})'
    }, inplace=True)
    
    with pd.ExcelWriter(buffer_excel, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Cuadro Comparativo', startrow=3)
        
        workbook = writer.book
        worksheet = writer.sheets['Cuadro Comparativo']
        
        # Estilos visuales
        HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="1E3A8A")
        SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
        HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        DATA_FONT = Font(name="Calibri", size=10)
        
        THIN_BORDER = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Título del Reporte
        worksheet['A1'] = "ENAEX - CUADRO COMPARATIVO DE OFERTAS"
        worksheet['A1'].font = TITLE_FONT
        worksheet['A2'] = f"Fecha de informe: {date.today().strftime('%d/%m/%Y')} | Moneda base: {moneda_vista}"
        worksheet['A2'].font = SUBTITLE_FONT
        
        # Formato de Encabezados de Tabla (Fila 4)
        for col_num in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
        
        # Formato de Filas de Datos
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=5, max_row=4 + len(df_export), min_col=1, max_col=len(df_export.columns)), start=5):
            use_zebra = (row_idx % 2 == 0)
            for cell in row:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if use_zebra:
                    cell.fill = ZEBRA_FILL
                
                col_header = worksheet.cell(row=4, column=cell.column).value
                if col_header in ['Precio Unitario', f'Total ({moneda_vista})']:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_header in ['Cantidad', 'Pos', 'Días para Entrega']:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_header in ['SOLPED', 'Moneda', 'UM', 'Centro']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Autoajuste inteligente de ancho de columnas
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4: continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return buffer_excel.getvalue()

def clean_str_pdf(txt):
    """Limpia caracteres especiales para evitar errores de codificación en PDF"""
    s = str(txt or '')
    reemplazos = {'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','á':'a','é':'e','í':'i','ó':'o','ú':'u','Ñ':'N','ñ':'n','°':''}
    for k, v in reemplazos.items():
        s = s.replace(k, v)
    return s.encode('latin-1', 'ignore').decode('latin-1')

class PDFReport(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 14)
        self.set_text_color(30, 58, 138)
        self.cell(0, 8, 'ENAEX - EVALUACION COMPARATIVA DE OFERTAS', ln=True, align='C')
        self.set_font('Helvetica', 'I', 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, f'Fecha de Emision: {date.today().strftime("%d/%m/%Y")}', ln=True, align='C')
        self.ln(4)

    def footer(self):
        self.set_y(-12)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Pagina {self.page_no()}/{{nb}} - Documento Generado Automáticamente', align='C')

def generar_pdf(df, moneda_vista):
    """Genera un archivo PDF ejecutivo en formato horizontal (A4)"""
    pdf = PDFReport(orientation='L', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # Resumen superior
    monto_total = df["Monto Total Visualizado"].sum()
    pdf.set_fill_color(243, 244, 246)
    pdf.rect(10, pdf.get_y(), 277, 10, style='F')
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_text_color(30, 58, 138)
    pdf.cell(0, 8, f'  RESUMEN GENERAL: Total Ofertas Evaluadas: {len(df)}    |    Monto Acumulado ({moneda_vista}): ${monto_total:,.2f}', ln=True)
    pdf.ln(4)

    # Encabezados de tabla PDF
    cols = [
        ("SOLPED", 30),
        ("Material", 80),
        ("Proveedor", 55),
        ("Cant.", 18),
        ("Mon", 18),
        (f"Total ({moneda_vista})", 40),
        ("Entrega", 36)
    ]

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(30, 58, 138)
    pdf.set_text_color(255, 255, 255)

    for name, width in cols:
        pdf.cell(width, 7, name, border=1, align='C', fill=True)
    pdf.ln()

    # Filas de datos PDF
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(0, 0, 0)
    
    fill = False
    for _, row in df.iterrows():
        if pdf.get_y() > 180:
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_fill_color(30, 58, 138)
            pdf.set_text_color(255, 255, 255)
            for name, width in cols:
                pdf.cell(width, 7, name, border=1, align='C', fill=True)
            pdf.ln()
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(0, 0, 0)

        pdf.set_fill_color(249, 250, 251) if fill else pdf.set_fill_color(255, 255, 255)

        solped = clean_str_pdf(row.get('SOLPED', ''))[:18]
        material = clean_str_pdf(row.get('Material', ''))[:48]
        proveedor = clean_str_pdf(row.get('Proveedor Visual', ''))[:32]
        cant = f"{row.get('Cantidad', 0):,.0f}"
        mon = clean_str_pdf(row.get('Moneda', 'CLP'))
        monto = f"${row.get('Monto Total Visualizado', 0):,.2f}"
        dias = f"{int(row.get('Días para Entrega', 0))} dias"

        pdf.cell(cols[0][1], 6, solped, border=1, align='C', fill=True)
        pdf.cell(cols[1][1], 6, material, border=1, align='L', fill=True)
        pdf.cell(cols[2][1], 6, proveedor, border=1, align='L', fill=True)
        pdf.cell(cols[3][1], 6, cant, border=1, align='C', fill=True)
        pdf.cell(cols[4][1], 6, mon, border=1, align='C', fill=True)
        pdf.cell(cols[5][1], 6, monto, border=1, align='R', fill=True)
        pdf.cell(cols[6][1], 6, dias, border=1, align='C', fill=True)
        pdf.ln()
        fill = not fill

    return bytes(pdf.output())

# =============================================================================
# FUNCIONES AUXILIARES Y BÚSQUEDA ROBUSTA
# =============================================================================
def procesar_y_reparar_planilla(df):
    if df is None or df.empty:
        return df

    palabras_clave = ['sp', 'solped', 'material', 'pos', 'texto breve', 'centro', 'cantidad']
    header_idx = -1
    
    for idx in range(min(20, len(df))):
        row_values = [str(val).lower() for val in df.iloc[idx]]
        matches = sum(1 for val in row_values for kw in palabras_clave if kw in val)
        if matches >= 2:
            header_idx = idx
            break

    if header_idx != -1:
        nuevas_columnas = []
        for i, val in enumerate(df.iloc[header_idx]):
            val_str = str(val).strip()
            if val_str.lower() in ['nan', 'none', '']:
                col_orig = str(df.columns[i])
                if not col_orig.startswith("Unnamed"):
                    nuevas_columnas.append(col_orig)
                else:
                    nuevas_columnas.append(f"Col_Vacia_{i}")
            else:
                nuevas_columnas.append(val_str)
        
        df.columns = nuevas_columnas
        df = df.iloc[header_idx + 1:].reset_index(drop=True)
        
        if not df.empty:
            primer_col = df.columns[0]
            df = df[df[primer_col].astype(str).str.strip() != str(primer_col).strip()].reset_index(drop=True)

    mapeo_columnas = {
        'Unnamed: 6': 'UM', 'Unnamed: 7': 'Solicitante', 'Unnamed: 8': 'Centro',
        'Unnamed: 9': 'Tipo de posición', 'Unnamed: 10': 'G. compras', 'Unnamed: 11': 'Mod. el',
        'Unnamed: 12': 'Urgencia', 'Unnamed: 13': 'NS', 'Unnamed: 14': 'Contrato marco',
        'Unnamed: 15': 'Observación', 'Unnamed: 16': 'Responsable', 'Unnamed: 41': 'Total general'
    }
    df = df.rename(columns={k: v for k, v in mapeo_columnas.items() if k in df.columns})

    vistos = {}
    columnas_deduplicadas = []
    for c in df.columns:
        c_str = str(c).strip()
        if c_str in vistos:
            vistos[c_str] += 1
            columnas_deduplicadas.append(f"{c_str}_{vistos[c_str]}")
        else:
            vistos[c_str] = 0
            columnas_deduplicadas.append(c_str)
    df.columns = columnas_deduplicadas

    cols = list(df.columns)
    id_col = None
    
    for c in cols:
        if str(c).lower() in ['sp', 'solped', 'solicitud']:
            id_col = c
            break
            
    if not id_col:
        for c in cols:
            if any(kw in str(c).lower() for kw in ['sp', 'solped', 'solicitud', 'pr', 'requerimiento', 'pedido']):
                id_col = c
                break
                
    if id_col and id_col in cols:
        cols.remove(id_col)
        cols.insert(0, id_col)
        df = df[cols]

    df = df.dropna(how='all')
    return df

def extraer_materiales_de_masivo(df, id_solped):
    if df is None or df.empty:
        return []
        
    raw_search = str(id_solped).strip()
    if not raw_search or raw_search.lower() in ["(id solped)", "none", "nan"]:
        return []
        
    digits_search = re.sub(r'\D', '', raw_search)
    
    sp_cols = [c for c in df.columns if any(kw in str(c).lower() for kw in ['sp', 'solped', 'solicitud', 'pr', 'requerimiento', 'doc', 'pedido', 'compra'])]
    if not sp_cols:
        sp_cols = list(df.columns)

    df_filtrado = pd.DataFrame()
    
    for col in sp_cols:
        col_str = df[col].astype(str).str.strip()
        mask = col_str.str.lower() == raw_search.lower()
        
        if not mask.any() and digits_search:
            col_digits = col_str.apply(lambda x: re.sub(r'\D', '', str(x)))
            mask = col_digits == digits_search
            
        if not mask.any():
            mask = col_str.str.lower().str.contains(raw_search.lower(), regex=False)

        if mask.any():
            df_filtrado = df[mask]
            break

    if df_filtrado.empty:
        return []

    posiciones = []
    for idx, row in enumerate(df_filtrado.to_dict('records')):
        def get_val(keys, default):
            for k in keys:
                for col in row.keys():
                    if k in str(col).lower() and pd.notna(row[col]) and str(row[col]).strip() != "":
                        return row[col]
            return default

        # Función corregida para el parseo de números y formato chileno/latino
        def clean_num(val, default=0.0):
            try:
                if isinstance(val, (int, float)): 
                    return float(val)
                s = str(val).strip()
                s = re.sub(r'[^\d.,-]', '', s)
                if '.' in s and ',' in s:
                    s = s.replace('.', '').replace(',', '.')
                elif '.' in s and len(s.split('.')[-1]) == 3:
                    s = s.replace('.', '')
                elif ',' in s:
                    s = s.replace(',', '.')
                return float(s)
            except Exception:
                return default

        posiciones.append({
            "Pos": int(idx + 1),
            "Material": str(get_val(['texto', 'desc', 'material', 'denominacion', 'item', 'artículo', 'articulo', 'breve'], f"Material {idx+1}")),
            "Centro": str(get_val(['centro', 'plant', 'almacen', 'alm'], "E001")),
            "Cantidad": clean_num(get_val(['cant', 'cantidad', 'ctd'], 1.0), 1.0),
            "UM": str(get_val(['um', 'unidad', 'unid', 'medida'], "C/U")).upper(),
            "Precio Unitario": clean_num(get_val(['precio', 'monto', 'val', 'costo', 'p.u', 'neto'], 0.0), 0.0),
            "Moneda": str(get_val(['moneda', 'curr', 'mon'], "CLP")).upper(),
            "Proveedor": str(get_val(['proveedor', 'vendor', 'prov', 'nam'], "")),
            "Calendario de entrega": date.today(),
            "Observaciones": str(get_val(['obs', 'observacion', 'comentario'], ""))
        })
        
    return posiciones

def convertir_moneda(monto, moneda_origen, tc_usd, tc_uf, tc_eur):
    """Calcula importes equivalentes en CLP, USD y EUR"""
    monto = float(monto or 0.0)
    moneda_origen = str(moneda_origen).upper()
    
    if moneda_origen == "CLP":
        clp = monto
    elif moneda_origen == "USD":
        clp = monto * tc_usd
    elif moneda_origen == "UF":
        clp = monto * tc_uf
    elif moneda_origen == "EUR":
        clp = monto * tc_eur
    else:
        clp = monto
        
    usd = clp / tc_usd if tc_usd > 0 else 0.0
    eur = clp / tc_eur if tc_eur > 0 else 0.0
    return clp, usd, eur

# =============================================================================
# INICIALIZACIÓN DE ESTADO
# =============================================================================
if "df_masivo" not in st.session_state:
    st.session_state.df_masivo = None
if "ofertas_manuales" not in st.session_state:
    st.session_state.ofertas_manuales = []

# =============================================================================
# ENCABEZADO Y PARÁMETROS GLOBALES
# =============================================================================
st.markdown("<div class='main-header'>⚡ Sistema Integrado de Evaluación de Ofertas</div>", unsafe_allow_html=True)

with st.sidebar:
    st.header("⚙️ Parámetros de Cambio")
    
    indicadores = obtener_indicadores_tiempo_real()
    
    if indicadores["estado"]:
        st.success("🟢 Indicadores actualizados en tiempo real")
    else:
        st.warning("⚠️ Sin conexión a API. Usando valores por defecto.")
        
    if st.button("🔄 Actualizar Tasas API", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.divider()

    tc_usd = st.number_input("Tipo de Cambio USD / CLP", value=indicadores["USD"], step=1.0, format="%.2f")
    tc_uf = st.number_input("Tipo de Cambio UF / CLP", value=indicadores["UF"], step=100.0, format="%.2f")
    tc_eur = st.number_input("Tipo de Cambio EUR / CLP", value=indicadores["EUR"], step=1.0, format="%.2f")
    st.divider()
    
    st.header("📂 Carga de Archivo Base")
    file_masivo = st.file_uploader("Cargar Planilla Maestro/SOLPEDs (Excel/CSV)", type=["xlsx", "xls", "csv", "xlsm"])
    
    if file_masivo:
        try:
            if file_masivo.name.endswith(".csv"):
                df_raw = pd.read_csv(file_masivo)
            else:
                dict_dfs = pd.read_excel(file_masivo, sheet_name=None, engine='openpyxl')
                df_raw = pd.concat(dict_dfs.values(), ignore_index=True)
                
            df_clean = df_raw.dropna(axis=1, how='all').dropna(axis=0, how='all')
            df_procesado = procesar_y_reparar_planilla(df_clean)
            
            df_procesado['cantidad_nulos'] = df_procesado.isnull().sum(axis=1)
            df_procesado = df_procesado.sort_values(by='cantidad_nulos').drop(columns=['cantidad_nulos']).reset_index(drop=True)
            
            st.session_state.df_masivo = df_procesado
            st.success(f"Planilla cargada y limpiada correctamente ({len(st.session_state.df_masivo)} filas)")
        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")

# =============================================================================
# VISTA PREVIA DE DATOS CARGADOS
# =============================================================================
if st.session_state.df_masivo is not None:
    with st.expander("👀 Vista Previa de la Planilla Base Cargada", expanded=False):
        st.write("Mostrando los datos procesados. La columna 'SP' ha sido priorizada en la primera posición para fácil lectura.")
        st.dataframe(st.session_state.df_masivo, use_container_width=True)

tabs = st.tabs(["✏️ Evaluación por SOLPED", "➕ Carga Manual / Directa", "📊 Cuadro Comparativo Integrado"])

# =============================================================================
# TAB 1: EVALUACIÓN POR SOLPED (AUTOGESTIÓN)
# =============================================================================
with tabs[0]:
    st.subheader("✏️ Evaluación por SOLPED")
    
    col_input, col_btn = st.columns([3, 1])
    with col_input:
        solped_id = st.text_input("Buscar ID SOLPED en la planilla:", placeholder="Ej: PR175798 o 175798")
    with col_btn:
        st.write("")
        st.write("")
        btn_extraer = st.button("📤 Extraer Materiales", type="primary", use_container_width=True)

    if (btn_extraer or solped_id) and solped_id.strip():
        if st.session_state.df_masivo is not None:
            materiales = extraer_materiales_de_masivo(st.session_state.df_masivo, solped_id)
            if materiales:
                st.session_state[f"editor_{solped_id}"] = pd.DataFrame(materiales)
                st.success(f"Se encontraron {len(materiales)} posiciones para la SOLPED **{solped_id}**")
            else:
                st.warning(f"No se encontraron registros para la SOLPED '{solped_id}'. Verifica si fue cargada en el panel lateral.")
        else:
            st.info("Carga una planilla maestra en el menú lateral para realizar la búsqueda automática por SOLPED.")

    key_editor = f"editor_{solped_id}" if (solped_id and f"editor_{solped_id}" in st.session_state) else "editor_default"
    
    df_inicial = st.session_state.get(key_editor, pd.DataFrame([{
        "Pos": 1, "Material": "(Material)", "Centro": "(Centro)", "Cantidad": 1.0, 
        "UM": "C/U", "Precio Unitario": 0.0, "Moneda": "CLP", 
        "Proveedor": "", "Calendario de entrega": date.today(), "Observaciones": ""
    }]))

    if not df_inicial.empty:
        df_inicial["Precio Unitario"] = pd.to_numeric(df_inicial["Precio Unitario"], errors='coerce').fillna(0.0)
        df_inicial["Cantidad"] = pd.to_numeric(df_inicial["Cantidad"], errors='coerce').fillna(1.0)
        df_inicial["Calendario de entrega"] = pd.to_datetime(df_inicial["Calendario de entrega"]).dt.date

    edited_df = st.data_editor(
        df_inicial,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Pos": st.column_config.NumberColumn("Pos", disabled=True),
            "Precio Unitario": st.column_config.NumberColumn("Precio Unitario", format="$ %.2f"),
            "Moneda": st.column_config.SelectboxColumn("Moneda", options=["CLP", "USD", "EUR"]),
            "Calendario de entrega": st.column_config.DateColumn("Fecha Entrega")
        }
    )

    if st.button("💾 Guardar Oferta de SOLPED en Comparativo", type="primary"):
        registros = edited_df.to_dict('records')
        for r in registros:
            clp, usd, eur = convertir_moneda(r["Precio Unitario"] * r["Cantidad"], r["Moneda"], tc_usd, tc_uf, tc_eur)
            r["SOLPED"] = solped_id if solped_id else "N/A"
            r["Total CLP"] = clp
            r["Total USD"] = usd
            r["Total EUR"] = eur
            st.session_state.ofertas_manuales.append(r)
        st.success("¡Oferta guardada exitosamente en el Cuadro Comparativo!")

# =============================================================================
# TAB 2: CARGA MANUAL INTEGRA / EDICIÓN DIRECTA POR SOLPED
# =============================================================================
with tabs[1]:
    st.subheader("➕ Carga Manual de Oferta Paso a Paso")
    
    col_s1, col_s2 = st.columns([3, 1])
    with col_s1:
        manual_solped = st.text_input("Ingresar N° SOLPED para Autocompletar:", placeholder="Ej: PR175798")
    with col_s2:
        st.write("")
        st.write("")
        btn_cargar_manual = st.button("📥 Cargar Requerimiento", use_container_width=True)

    if btn_cargar_manual and manual_solped:
        if st.session_state.df_masivo is not None:
            mats = extraer_materiales_de_masivo(st.session_state.df_masivo, manual_solped)
            if mats:
                st.session_state["manual_grid_df"] = pd.DataFrame(mats)
                st.success(f"Materiales cargados automáticamente desde la SOLPED {manual_solped}")
            else:
                st.warning(f"No se encontró la SOLPED {manual_solped} en el archivo base.")
        else:
            st.info("Sube una planilla en la barra lateral para autocompletar posiciones por SOLPED.")

    if "manual_grid_df" not in st.session_state:
        st.session_state["manual_grid_df"] = pd.DataFrame([{
            "Pos": 1, "Material": "Ítem Manual", "Cantidad": 1.0, "UM": "C/U",
            "Precio Unitario": 0.0, "Moneda": "CLP", "Proveedor": "", 
            "Calendario de entrega": date.today(), "Observaciones": ""
        }])

    df_manual = st.session_state["manual_grid_df"]
    if not df_manual.empty:
        df_manual["Precio Unitario"] = pd.to_numeric(df_manual["Precio Unitario"], errors='coerce').fillna(0.0)
        df_manual["Cantidad"] = pd.to_numeric(df_manual["Cantidad"], errors='coerce').fillna(1.0)
        df_manual["Calendario de entrega"] = pd.to_datetime(df_manual["Calendario de entrega"]).dt.date

    st.write("### Tabla de Cotización de Proveedor")
    
    cotizacion_df = st.data_editor(
        df_manual,
        num_rows="dynamic",
        use_container_width=True,
        key="cotizacion_manual_editor",
        column_config={
            "Precio Unitario": st.column_config.NumberColumn("Precio Unitario", format="$ %.2f"),
            "Moneda": st.column_config.SelectboxColumn("Moneda", options=["CLP", "USD", "EUR"]),
            "Calendario de entrega": st.column_config.DateColumn("Calendario de entrega")
        }
    )

    if st.button("💾 Guardar Cotización Manual Completa", type="primary"):
        items = cotizacion_df.to_dict('records')
        for item in items:
            clp, usd, eur = convertir_moneda(item["Precio Unitario"] * item["Cantidad"], item["Moneda"], tc_usd, tc_uf, tc_eur)
            item["SOLPED"] = manual_solped if manual_solped else "MANUAL"
            item["Total CLP"] = clp
            item["Total USD"] = usd
            item["Total EUR"] = eur
            st.session_state.ofertas_manuales.append(item)
        st.success("¡Cotización agregada al Cuadro Comparativo!")

# =============================================================================
# TAB 3: CUADRO COMPARATIVO INTEGRADO & DESCARGAS
# =============================================================================
with tabs[2]:
    st.subheader("📊 Cuadro Comparativo Integrado")
    
    if st.session_state.ofertas_manuales:
        df_comp = pd.DataFrame(st.session_state.ofertas_manuales)
        
        # Selector de Moneda de Visualización
        moneda_vista = st.radio(
            "💱 Seleccionar Moneda de Visualización:", 
            options=["CLP", "USD", "EUR"], 
            horizontal=True
        )
        
        # Asignar la columna total dinámica
        if moneda_vista == "CLP":
            df_comp["Monto Total Visualizado"] = df_comp["Total CLP"]
        elif moneda_vista == "USD":
            df_comp["Monto Total Visualizado"] = df_comp["Total USD"]
        elif moneda_vista == "EUR":
            df_comp["Monto Total Visualizado"] = df_comp["Total EUR"]

        # Calcular días restantes de entrega
        df_comp['Calendario de entrega'] = pd.to_datetime(df_comp['Calendario de entrega'])
        hoy = pd.Timestamp(date.today())
        df_comp['Días para Entrega'] = (df_comp['Calendario de entrega'] - hoy).dt.days
        df_comp['Días para Entrega'] = df_comp['Días para Entrega'].apply(lambda x: x if x > 0 else 0)
        
        df_comp['Proveedor Visual'] = df_comp['Proveedor'].replace("", "Sin Especificar")

        # Motor de recomendación visual
        st.markdown("### 🏆 Motor de Recomendación")
        st.info("💡 **Guía de colores:** Se resalta en **verde** la opción más económica y en **azul** la entrega más rápida para cada material.")
        
        def highlight_best(df):
            styles = pd.DataFrame('', index=df.index, columns=df.columns)
            for name, group in df.groupby(['SOLPED', 'Material']):
                if len(group) > 1:
                    min_monto_idx = group['Monto Total Visualizado'].idxmin()
                    min_dias_idx = group['Días para Entrega'].idxmin()
                    styles.loc[min_monto_idx, 'Monto Total Visualizado'] = 'background-color: #D1FAE5; color: #065F46; font-weight: bold;'
                    styles.loc[min_dias_idx, 'Días para Entrega'] = 'background-color: #DBEAFE; color: #1E3A8A; font-weight: bold;'
            return styles

        styled_df_comp = df_comp.style.apply(highlight_best, axis=None).format({
            "Monto Total Visualizado": "$ {:,.2f}",
            "Precio Unitario": "$ {:,.2f}",
            "Total CLP": "$ {:,.2f}",
            "Total USD": "$ {:,.2f}",
            "Total EUR": "$ {:,.2f}"
        })

        st.dataframe(styled_df_comp, use_container_width=True)
        
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            st.metric("Total Ofertas Registradas", len(df_comp))
        with col_c2:
            monto_acumulado = df_comp["Monto Total Visualizado"].sum()
            st.metric(f"Monto Total Acumulado ({moneda_vista})", f"$ {monto_acumulado:,.2f}")
            
        st.divider()
        st.subheader("📈 Gráficos Comparativos por SOLPED")
        
        col_graf1, col_graf2 = st.columns(2)
        
        with col_graf1:
            st.markdown(f"**💰 Comparativa de Monto Total por SOLPED ({moneda_vista})**")
            df_monto_solped = df_comp.groupby("SOLPED")["Monto Total Visualizado"].sum().reset_index()
            st.bar_chart(df_monto_solped, x="SOLPED", y="Monto Total Visualizado", color="SOLPED", height=350)
            
        with col_graf2:
            st.markdown("**⏳ Promedio Días de Entrega por SOLPED**")
            df_dias_solped = df_comp.groupby("SOLPED")["Días para Entrega"].mean().reset_index()
            st.bar_chart(df_dias_solped, x="SOLPED", y="Días para Entrega", color="SOLPED", height=350)

        st.divider()
        st.subheader("📥 Exportar Reportes")
        st.write("Genera y descarga el informe en tu formato de preferencia:")
        
        # Generar archivos
        bytes_excel = generar_excel_estilizado(df_comp, moneda_vista)
        bytes_pdf = generar_pdf(df_comp, moneda_vista)
        
        col_down1, col_down2, col_down3 = st.columns([1, 1, 2])
        
        with col_down1:
            st.download_button(
                label="📊 Descargar Excel Estilizado",
                data=bytes_excel,
                file_name=f"Reporte_Comparativo_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            
        with col_down2:
            st.download_button(
                label="📄 Descargar Reporte PDF",
                data=bytes_pdf,
                file_name=f"Reporte_Comparativo_{date.today()}.pdf",
                mime="application/pdf",
                use_container_width=True
            )

        st.write("")
        if st.button("🗑️ Limpiar Cuadro Comparativo", use_container_width=False):
            st.session_state.ofertas_manuales = []
            st.rerun()
    else:
        st.info("Aún no hay ofertas registradas en el Cuadro Comparativo.")
